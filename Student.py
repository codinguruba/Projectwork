import openpyxl

class StudentManager:


    def add_student(self,id,name,age):
        file_name = "students.xlsx"

        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active

        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["id","name","age"])

        sheet.append([id,name,age])
        print("Student data updated....")
        workbook.save(file_name)

    def read_student(self):
        file_name = "students.xlsx"

        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):
                print(f"ID: {row[0]}, Name: {row[1]}, Age: {row[2]}")


        except FileNotFoundError:
            print(f"{file_name} not found.")


    def delete_student(self, student_id):
        file_name = "students.xlsx"
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active

        # Iterate and delete the row found with matching student ID
            for row in sheet.iter_rows(min_row=2): 
                if row[0].value == student_id:
                    sheet.delete_rows(row[0].row)
                    workbook.save(file_name)
                    print(f"Student with ID {student_id} deleted.")
                
        
        except FileNotFoundError:
            print(f"{file_name} not found.")

